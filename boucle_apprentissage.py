#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sun Jun 18 18:23:29 2017

@author: majdagoumi
"""
from algo_MC_2 import *
from prise_2_decision import *
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border, Side, Alignment
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from load_object import *
from prise_3_decision import *
###

def f(values):
	return values.index(max(values))

def apprentissage(n):
    victoire = 0
    for k in range(n):
        victoire = manche()
    print("My policy : ",mypolicy)
    print("Stats victoires : ",stat_gain)
    print("Nb etat joué et nb etats ganés : ", nombre_etats_joue,nombre_etats_gagnes)
    print("Pourcentage de victoire : ",100*nombre_etats_gagnes[0]/nombre_etats_joue[0])

    result = [[f(values) for values in i]for i in mypolicy]
#    print("     ↑ low enemy card   | high enemy card ↓")
#    print("     <- low value hand  |  high value hand ->")
#    print('\n'.join([''.join(['{:4}'.format(item) for item in row]) for row in result]))

def apprentissage2(n,bet):
    victoire = 0
    for k in range(n):
        victoire = manche2(bet)
        if(k%100000 == 0):
            print("progres : ", (k/n)*100 ,"%")
    print("Statistiques :\n")
    print(victoire)
    print("Epsilon final, alpha : ",epsilon,", ",alpha)
    
def apprentissage3(n):
    victoires = 0
    defeats = 0
    gain = 0
    for k in range(n):
        result, player_statesActions,bool_as_choice,bool_pair, enemy_state,position_as, enemy_statesActions,enemy_bool_as_choice, enemy_bool_pair, player_state, enemy_position_as = symetricLearning()
        ###MAJ des stats : 
        #pourcentage_victoire, gain = update_stats_gains(player_statesActions,result,enemy_state,player_state)
        ###MAJ de mypolicy
        #print("Main du joueur : ", player_hand)
        update_value3(player_statesActions,bool_as_choice,bool_pair, enemy_state-1,result,position_as)
        update_value3(enemy_statesActions,enemy_bool_as_choice,enemy_bool_pair, player_state-1, - result,enemy_position_as)
        #victoire=update_stats_gains(player_statesActions,result,enemy_state,player_state)
        ###Fin de la MAJ de mypolicy
        if(k%100000 == 0):
            print("progres : ", (k/n)*100 ,"%")      
        gain += result
        if result > 0:
            victoires += 1
        elif result < 0 :
            defeats +=1
        if(k%100000 == 0):
            print("Progrès : ", int((k/n)*100) ,"%")  
    print("Pourcentage de victoires : ", int((victoires/n)*100000)/1000, "%")
    print("Pourcentage de victoires de l'adversaire : ", int((defeats/n)*100000)/1000, "%")
    print("Gain par euro investi : ", int((gain/n)*1000)/1000)
    

def f(k):
        return(0.05)
    

   
def save_mypolicy(p1,p2,p3): #Le fichier sauvegarde est un vecteur comportant les trois tableaux
#from load_object import *
#IND : Pour sauvegarder : ecrire save_mypolicy(policy_simple,policy_as,policy_pair)
#      Pour charger ecrire : policy_simple,policy_as,policy_pair = getPolicy()[0],getPolicy()[1],getPolicy()[2]

    policy = [p1,p2,p3]
    with open("mypolicy_IAvsIA_50M_v2", "wb") as file_handler:
        pickle.dump(policy,file_handler)
    
    

def save_to_xlsx():
    wb = Workbook()

    
    # grab the active worksheet
    ws = wb.active
    
    
                        
    
    ws['A1']="Simple :"
    ws['A2']=""
    ws.append(["","< 9"]+list(range(9,22))+["> 21"])
#########    
    policy = getPolicy()[:]
#########    
    i=1
    for row in policy[0][0] :
        for k in range(len(row)):
            row[k] = row[k].index(max(row[k]))
        if (i==1):
            ws.append(['A']+row)
        else :
            ws.append([i]+row)
        i+=1
    
    ws['A14'] = ""
    ws['A15'] = "As :"
    ws['A16'] = ""
    ws.append([""]+["A,"+str(k) for k in range (2,11)])
    
    i=1
    for row in policy[1][0] :
        for k in range(len(row)):
            row[k] = row[k].index(max(row[k]))
        if (i==1):
            ws.append(['A']+row)
        else :
            ws.append([i]+row)
        i+=1
    
    ws['A28'] = ""
    ws['A29'] = "Paires :"
    ws['A30'] = ""
    ws.append([""]+["A,A"]+[str(k)+","+str(k) for k in range (2,11)])
    
    i=1
    for row in policy[2][0] :
        for k in range(len(row)):
            row[k] = row[k].index(max(row[k]))
        if (i==1):
            ws.append(['A']+row)
        else :
            ws.append([i]+row)
        i+=1
    

    
    redFill = PatternFill(start_color='FF6666',
                   end_color='FF6666',
                   fill_type='solid')
    
    orangeFill = PatternFill(start_color='FFC966',
                        end_color='FFC966',
                        fill_type='solid')
    
    greenFill = PatternFill(start_color='6DC066',
                        end_color='6DC066',
                        fill_type='solid')
                        
    purpleFill = PatternFill(start_color='8067A2',
                        end_color='8067A2',
                        fill_type='solid')
    
    alignment=Alignment(horizontal='center',
                        vertical='center',
                        text_rotation=0,
                        wrap_text=False,
                        shrink_to_fit=False,
                        indent=0)
    
    
    rows = ws['A3:P13']
    side = Side(border_style='thin', color="FF000000")

    rows = list(rows)  # we convert iterator to list for simplicity, but it's not memory efficient solution
    max_y = len(rows) - 1  # index of the last row
    for pos_y, cells in enumerate(rows):
        max_x = len(cells) - 1  # index of the last cell
        for pos_x, cell in enumerate(cells):
            border = Border(
                left=cell.border.left,
                right=cell.border.right,
                top=cell.border.top,
                bottom=cell.border.bottom
            )
            border.left = side
            border.right = side
            border.top = side
            border.bottom = side

            if (cell.value==0 and pos_x!=0) :
                cell.fill=redFill
            if (cell.value==1 and pos_x!=0) :
                cell.fill=orangeFill
            if (cell.value==2 and pos_x!=0) :
                cell.fill=greenFill
            if (cell.value==3 and pos_x!=0) : 
                cell.style='Accent4'
            cell.border = border
            cell.alignment=alignment
    
    rows = ws['A17:J27']
    side = Side(border_style='thin', color="FF000000")

    rows = list(rows)  # we convert iterator to list for simplicity, but it's not memory efficient solution
    max_y = len(rows) - 1  # index of the last row
    for pos_y, cells in enumerate(rows):
        max_x = len(cells) - 1  # index of the last cell
        for pos_x, cell in enumerate(cells):
            border = Border(
                left=cell.border.left,
                right=cell.border.right,
                top=cell.border.top,
                bottom=cell.border.bottom
            )
            border.left = side
            border.right = side
            border.top = side
            border.bottom = side

            if (cell.value==0 and pos_x!=0) :
                cell.fill=redFill
            if (cell.value==1 and pos_x!=0) :
                cell.fill=orangeFill
            if (cell.value==2 and pos_x!=0) :
                cell.fill=greenFill
            if (cell.value==3 and pos_x!=0) : 
                cell.style='Accent4'
            cell.border = border
            cell.alignment=alignment
    
    rows = ws['A31:K41']
    side = Side(border_style='thin', color="FF000000")

    rows = list(rows)  # we convert iterator to list for simplicity, but it's not memory efficient solution
    max_y = len(rows) - 1  # index of the last row
    for pos_y, cells in enumerate(rows):
        max_x = len(cells) - 1  # index of the last cell
        for pos_x, cell in enumerate(cells):
            border = Border(
                left=cell.border.left,
                right=cell.border.right,
                top=cell.border.top,
                bottom=cell.border.bottom
            )
            border.left = side
            border.right = side
            border.top = side
            border.bottom = side
            if (cell.value==0 and pos_x!=0) :
                cell.fill=redFill
            if (cell.value==1 and pos_x!=0) :
                cell.fill=orangeFill
            if (cell.value==2 and pos_x!=0) :
                cell.fill=greenFill
            if (cell.value==3 and pos_x!=0) : 
                cell.fill=purpleFill
            cell.border = border
            cell.alignment=alignment
    
    # Save the file
    wb.save("mypolicy_IAvsIA_50M_v2.xlsx")



def presenter_resultats(mypolicy):
    matrice_simple = mypolicy[0]
    matrice_as = mypolicy[1]
    matrice_paire = [2]

    
def graphe_vitesse_apprentissage(a_min=0, pas=2 * 10**5,a_max = 10**7, log=False,
                                 nb_test=2*10**5):

    from matplotlib.pyplot import plot, show, xlabel, ylabel, figure

    
    lx = [1]
    ly = [test_sans_apprendre(nb_test)]

    apprentissage2(a_min, 1, lambda k: 0.05)
    x = a_min
    
    while x < a_max :
        print ("iteration ", x, "sur ", a_max)
        ly.append(test_sans_apprendre(nb_test))
        lx.append(x)
        
        if log:
            apprentissage2(x * (pas - 1), 1, lambda k: 0.05)
            x *= pas
        else :
            apprentissage2(pas, 1, lambda k: 0.05)
            x += pas
                
        

    print (lx, ly)
    fig = figure()
    ax = fig.add_subplot(2, 1, 1)
    ax.set_xscale('log')
    xlabel("nombre d'iterations d'apprentissage")
    ylabel("gain moyen pour une mise de 1")
    plot(lx, ly)
    show()

    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    